import io
import os
import fitz
import pypdf
import base64
import logging
import subprocess
import requests

from PIL import Image
from docx import Document
from langchain.text_splitter import CharacterTextSplitter

from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter


extract_prompt = f'''
You are given a page from a document - could be a PDF, PowerPoint, Excel or Word document. 
Your task is to extract all relevant information from it. It is important to keep the extract accurate, comprehensive, and well-structured, as the extracted information will be used as input for a Retrieval-Augmented Generation (RAG) model.

Please follow these detailed instructions:

1. **Identify and Extract Content**:
   - **Chart**: If you identify a chart, provide a detailed description of its content. Include:
     - The type of chart (e.g., bar, line, pie, etc.).
     - The axes labels and their units.
     - The data points or trends observed.
     - A summary of what the chart depicts.
   
   - **Table**: If there is a table present, extract it while preserving the original table formatting. Ensure:
     - All headers and sub-headers are correctly identified.
     - The hierarchical structure of the table is maintained.
     - Any notable data points or trends are highlighted.
   
   - **Text**: If the object is neither a chart nor a table, extract the remaining text as it is, i.e. keeping the original text unchanged. Also, ensure:
     - All bullet points, numbering, and indentations are preserved.
     - Any emphasized text (bold, italics, etc.) is accurately represented.
     
2. **Summarize the Slide**:
   - Provide a summarized description of the slide, capturing the main points and purpose of the slide.
'''

formatting = '''
**Formatting re-enabled**:
Where appropriate, improve the readability by formatting the response using Markdown, e.g.:
- **bold**
- _italics_
- list
- tables
- header tags (start from ###)
'''


def convert_office_to_pdf(input_file, output_file, folder_pref, environ):
    # Using LibreOffice to convert pptx to pdf via command line
    # LibreOffice headless mode is used for running without GUI
    import time
    
    output_dir = os.path.dirname(output_file)
    command = ['libreoffice', '--headless', '--convert-to', 'pdf', input_file, '--outdir', output_dir]

    subprocess.run(command, env=environ, check=True)
    
def convert_office_to_pdf_optimized(input_file, output_file, folder_pref, environ):
    """
    Convert Office files to PDF with aggressive process cleanup
    """
    import subprocess
    import os
    import time
    
    output_dir = os.path.dirname(output_file)
    
    # Add additional flags to prevent process persistence
    command = [
        'libreoffice', 
        '--headless',
        '--convert-to', 'pdf',
        '--norestore',
        '--nolockcheck',
        '--nologo',
        '--nodefault',
        '--nofirststartwizard',
        input_file, 
        '--outdir', output_dir
    ]
    
    # Run conversion with timeout
    process = None
    try:
        process = subprocess.Popen(
            command, 
            env=environ,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            preexec_fn=os.setsid  # Create new process group
        )
        
        # Wait for completion with timeout
        stdout, stderr = process.communicate(timeout=120)
        
        # Check if conversion was successful
        if process.returncode != 0:
            error_msg = stderr.decode('utf-8', errors='ignore') if stderr else 'Unknown error'
            raise subprocess.CalledProcessError(
                process.returncode, command, stdout, stderr
            )
        
        # Verify output file was created
        if not os.path.exists(output_file):
            raise FileNotFoundError(
                "Conversion completed but output file not found: {}".format(output_file)
            )
        
        # Small delay to ensure file is fully written
        time.sleep(0.5)
        
    except subprocess.TimeoutExpired:
        # Kill the entire process group if timeout
        if process:
            try:
                os.killpg(os.getpgid(process.pid), 9)
            except:
                pass
        raise Exception("LibreOffice conversion timeout after 120 seconds")
    
    except Exception as e:
        # Kill process on any error
        if process and process.poll() is None:
            try:
                os.killpg(os.getpgid(process.pid), 9)
            except:
                pass
        raise
    
    finally:
        # Ensure process is terminated
        if process and process.poll() is None:
            try:
                process.terminate()
                process.wait(timeout=5)
            except:
                try:
                    process.kill()
                except:
                    pass
        
        # Kill any lingering soffice processes from this conversion
        try:
            subprocess.run(
                ['pkill', '-9', '-f', 'soffice.*{}'.format(os.path.basename(input_file))],
                stderr=subprocess.DEVNULL,
                timeout=2
            )
        except:
            pass

def pdf_to_base64(pdf_path, page_num, zoom=1.5):
    document = fitz.open(pdf_path)
    page = document.load_page(page_num)

    # Convert PDF page to a pixmap (image)
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))

    # Create an Image object from the pixmap (image)
    image = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)

    # Save the image to a bytes buffer
    buffered = io.BytesIO()
    image.save(buffered, format='JPEG')

    # Encode the image bytes to a base64 string
    base64_string = base64.b64encode(buffered.getvalue()).decode('utf-8')

    # Format the result with the appropriate prefix
    base64_string = f'data:image/jpeg;base64,{base64_string}'
    document.close()
    
    return base64_string


def pdf_to_base64_2(pdf_path, page_num, zoom=1.5):
    """Memory-optimized version"""
    document = fitz.open(pdf_path)
    try:
        page = document.load_page(page_num)
        
        # Convert directly to JPEG bytes without intermediate PIL Image
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        
        # Get JPEG bytes directly from pixmap
        jpeg_bytes = pix.tobytes(output="jpeg", jpg_quality=85)
        
        # Clear pixmap from memory
        del pix
        
        # Encode to base64
        base64_string = base64.b64encode(jpeg_bytes).decode('utf-8')
        
        # Clear jpeg bytes
        del jpeg_bytes
        
        # Format the result
        return f'data:image/jpeg;base64,{base64_string}'
        
    finally:
        document.close()

def extract_text_docx(file_path, dirname, basename, encoder):
    splitter = CharacterTextSplitter(chunk_size=1000, separator='\n', chunk_overlap=100)
    doc = Document(file_path)
    
    paragraphs = "\n".join([para.text for para in doc.paragraphs if para.text.strip() != ""])
    chunks = splitter.split_text(paragraphs)

    return [
        {
            'Dir_Name': dirname,
            'File_Name': basename,
            'Chunk_Id': i + 1,
            'Chunk_Text': text,
            'Chunk_Length': len(text),
            'Token_Count': len(encoder.encode(text))
        }
        for i, text in enumerate(chunks)
    ]


def extract_text_pdf(file_path, dirname, basename, encoder):
    chunks = []
    
    with open(file_path, 'rb') as file:
        reader = pypdf.PdfReader(file)

        for page_num in range(len(reader.pages)):
            page = reader.pages[page_num]
            text = page.extract_text()

            chunks.append({
                'Dir_Name': dirname,
                'File_Name': basename,
                'Chunk_Id': page_num + 1,
                'Chunk_Text': text,
                'Chunk_Length': len(text),
                'Token_Count': len(encoder.encode(text))
            })
            
    return chunks


def extract_image_pdf(file_path, dirname, basename, openai, model_name, ppmt):
    chunks = []
    price_total = 0
    tokens_total = 0
    
    with open(file_path, 'rb') as file:
        reader = pypdf.PdfReader(file)
        pages_total = len(reader.pages)
        
        for page_num in range(len(reader.pages)):
            base64_string = pdf_to_base64(file_path, page_num)

            image_data = [{
                'type': 'image_url',
                'image_url': {
                    'url': base64_string,
                }
            }]

            content = [{'type': 'text', 'text': extract_prompt}] + image_data
            
            response = openai.chat.completions.create(
                model=model_name,
                messages=[
                    {'role': 'developer', 'content': formatting},
                    {'role': 'user', 'content': content}
                ],
                reasoning_effort='minimal',
                verbosity='low'
                # temperature=0)
            )
               

            # print(f"The content of the page {page_num} from the file {file_path} successfully extracted.")
            logging.info(f"The content of the page {page_num} from the file {file_path} successfully extracted.")
            text = response.choices[0].message.content

            tokens_complet = response.usage.completion_tokens
            tokens_total += tokens_complet
            
            price_total += (ppmt[0] * response.usage.prompt_tokens + ppmt[1] * tokens_complet) / 1000000

            # if tokens_complet > 8192:
            #     print(f"The number of tokens from the page {page_num} of the file {file_path} exceeds the limit.")
            #     continue

            chunks.append({
                'Dir_Name': dirname,
                'File_Name': basename,
                'Chunk_Id': page_num + 1,
                'Chunk_Text': text,
                'Chunk_Length': len(text),
                'Token_Count': tokens_complet
            })
            
    
    return chunks, pages_total, tokens_total, price_total

def extract_image_pdf_2(file_path, dirname, basename, openai, model_name, ppmt):
    """
    Extract content from PDF with controlled concurrency and batching
    to prevent memory issues
    """
    import pypdf
    import logging
    import gc
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    chunks = []
    price_total = 0
    tokens_total = 0
    
    # Get total page count
    with open(file_path, 'rb') as file:
        reader = pypdf.PdfReader(file)
        pages_total = len(reader.pages)
    
    print("Processing {} with {} pages".format(basename, pages_total))
    
    def process_single_page(page_num):
        """Process a single page with proper memory cleanup"""
        base64_string = None
        response = None
        
        try:
            # Generate base64 for this page
            base64_string = pdf_to_base64(file_path, page_num)

            image_data = [{
                'type': 'image_url',
                'image_url': {
                    'url': base64_string,
                }
            }]

            content = [{'type': 'text', 'text': extract_prompt}] + image_data
            
            # Make API call
            response = openai.chat.completions.create(
                model=model_name,
                messages=[
                    {'role': 'developer', 'content': formatting},
                    {'role': 'user', 'content': content}
                ],
                reasoning_effort='minimal',
                verbosity='low'
            )

            logging.info("Page {} from {} successfully extracted".format(page_num, file_path))
            
            # Extract results
            text = response.choices[0].message.content
            tokens_complet = response.usage.completion_tokens
            page_price = (ppmt[0] * response.usage.prompt_tokens + ppmt[1] * tokens_complet) / 1000000

            page_chunk = {
                'Dir_Name': dirname,
                'File_Name': basename,
                'Chunk_Id': page_num + 1,
                'Chunk_Text': text,
                'Chunk_Length': len(text),
                'Token_Count': tokens_complet
            }
            
            return page_chunk, tokens_complet, page_price
            
        except Exception as e:
            logging.error("Error processing page {} in file {}: {}".format(page_num, file_path, e))
            # Return empty chunk on error
            error_chunk = {
                'Dir_Name': dirname,
                'File_Name': basename,
                'Chunk_Id': page_num + 1,
                'Chunk_Text': 'Error processing page: {}'.format(str(e)),
                'Chunk_Length': 0,
                'Token_Count': 0
            }
            return error_chunk, 0, 0
            
        finally:
            # Aggressive memory cleanup
            if base64_string:
                del base64_string
            if response:
                del response
            
            # Clear image_data and content references
            try:
                del image_data, content
            except:
                pass
            
            # Force garbage collection for this thread
            gc.collect()
    
    BATCH_SIZE = 20  # Process 20 pages at a time
    MAX_WORKERS = 3  
    
    for batch_start in range(0, pages_total, BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, pages_total)
        batch_pages = range(batch_start, batch_end)
        
        print("  Processing pages {}-{}/{}".format(batch_start + 1, batch_end, pages_total))
        
        # Process this batch with limited concurrency
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # Submit all page processing tasks for this batch
            future_to_page = {
                executor.submit(process_single_page, page_num): page_num 
                for page_num in batch_pages
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_page):
                page_num = future_to_page[future]
                try:
                    page_chunk, page_tokens, page_price = future.result()
                    chunks.append(page_chunk)
                    tokens_total += page_tokens
                    price_total += page_price
                    
                except Exception as e:
                    logging.error("Critical error processing page {} in {}: {}".format(
                        page_num, basename, e))
                    # Add error chunk to maintain page count
                    error_chunk = {
                        'Dir_Name': dirname,
                        'File_Name': basename,
                        'Chunk_Id': page_num + 1,
                        'Chunk_Text': 'Critical error: {}'.format(str(e)),
                        'Chunk_Length': 0,
                        'Token_Count': 0
                    }
                    chunks.append(error_chunk)
                
                finally:
                    # Cleanup after each completed future
                    del future
            
            # Final cleanup after batch
            del future_to_page
            gc.collect()
        
        # Progress update after each batch
        print("  Completed {}/{} pages. Tokens: {}, Cost: ${:.4f}".format(
            len(chunks), pages_total, tokens_total, price_total))
    
    # Sort chunks by page number to maintain order
    chunks.sort(key=lambda x: x['Chunk_Id'])
    
    print("Completed processing {}: {} pages, {} tokens, ${:.4f}".format(
        basename, len(chunks), tokens_total, price_total))
    
    return chunks, pages_total, tokens_total, price_total


def extract_image_pdf_3(file_path, dirname, basename, openai, model_name, ppmt):  # Works for both GPT-5 and GPT-4
    chunks = []
    price_total = 0
    tokens_total = 0
    
    with open(file_path, 'rb') as file:
        reader = pypdf.PdfReader(file)
        pages_total = len(reader.pages)
        
        for page_num in range(len(reader.pages)):
            base64_string = pdf_to_base64(file_path, page_num)

            image_data = [{
                'type': 'image_url',
                'image_url': {
                    'url': base64_string,
                }
            }]

            content = [{'type': 'text', 'text': extract_prompt}] + image_data
            
            # Prepare base parameters
            params = {
                'model': model_name,
                'messages': [
                    {'role': 'user', 'content': formatting},
                    {'role': 'user', 'content': content}
                ]
            }
            
            # Add GPT-5 specific parameters
            if 'gpt-5' in model_name.lower():
                params['reasoning_effort'] = 'minimal'
                params['verbosity'] = 'low'
                params['messages'][0]['role'] = 'developer'  # Use 'developer' role for GPT-5
            
            response = openai.chat.completions.create(**params)
               
            logging.info(f"The content of the page {page_num} from the file {file_path} successfully extracted.")
            text = response.choices[0].message.content

            tokens_complet = response.usage.completion_tokens
            tokens_total += tokens_complet
            
            price_total += (ppmt[0] * response.usage.prompt_tokens + ppmt[1] * tokens_complet) / 1000000

            chunks.append({
                'Dir_Name': dirname,
                'File_Name': basename,
                'Chunk_Id': page_num + 1,
                'Chunk_Text': text,
                'Chunk_Length': len(text),
                'Token_Count': tokens_complet
            })
            
    
    return chunks, pages_total, tokens_total, price_total



def extract_image_pdf_gemini(file_path, dirname, basename, gemini_config, model_name, ppmt):
    chunks = []
    price_total = 0
    tokens_total = 0

    # Extract endpoint and API key from config
    endpoint = gemini_config['endpoint']
    api_key = gemini_config['api_key']
    
    # Use correct Merck header format
    headers = {
        "Content-Type": "application/json",
        "X-Merck-APIKey": api_key
    }
    
    with open(file_path, 'rb') as file:
        reader = pypdf.PdfReader(file)
        pages_total = len(reader.pages)
        
        for page_num in range(len(reader.pages)):
            base64_string = pdf_to_base64(file_path, page_num)
            
            # Remove data URL prefix if present
            if base64_string.startswith('data:image'):
                base64_string = base64_string.split(',', 1)[1]
            
            # Merck Gemini API format with image
            payload = {
                "contents": {
                    "role": "user",
                    "parts": [
                        {"text": formatting + "\n\n" + extract_prompt},
                        {
                            "inline_data": {
                                "mime_type": "image/png",
                                "data": base64_string
                            }
                        }
                    ]
                },
                "generation_config": {
                    "temperature": 0,
                    "maxOutputTokens": 8192
                }
            }
            
            response = requests.post(
                endpoint,
                headers=headers,
                json=payload,
                timeout=120
            )
            
            response.raise_for_status()
            result = response.json()
            
            logging.info(f"The content of the page {page_num} from the file {file_path} successfully extracted.")
            
            # Extract text from response
            text = result['candidates'][0]['content']['parts'][0]['text']
            
            # Extract tokens
            usage = result.get('usageMetadata', {})
            tokens_prompt = usage.get('promptTokenCount', 0)
            tokens_complet = usage.get('candidatesTokenCount', 0)
            tokens_total += tokens_complet
            
            # Calculate price
            price_total += (ppmt[0] * tokens_prompt + ppmt[1] * tokens_complet) / 1000000

            chunks.append({
                'Dir_Name': dirname,
                'File_Name': basename,
                'Chunk_Id': page_num + 1,
                'Chunk_Text': text,
                'Chunk_Length': len(text),
                'Token_Count': tokens_complet
            })
    
    return chunks, pages_total, tokens_total, price_total

def prepare_excel_for_pdf(input_file, max_rows=1000):
    workbook = load_workbook(input_file)
    sheets_to_remove = []
    
    for sheet in workbook.worksheets:        
        if sheet.max_row > max_rows:
            sheets_to_remove.append(sheet.title)
        else:
            total_width, total_height = get_sheet_dimensions(sheet)

            if total_width > total_height:
                sheet.page_setup.orientation = 'landscape'
            else:
                sheet.page_setup.orientation = 'portrait'
            
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToHeight = total_width > total_height
            sheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)

    # Remove sheets marked for removal
    for sheet_name in sheets_to_remove:
        std = workbook[sheet_name]
        workbook.remove(std)
        
    # Save the modifications
    workbook.save(input_file)
    workbook.close()


def get_sheet_dimensions(sheet):
    total_width = 0
    total_height = 0

    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        total_width += 6 * (sheet.column_dimensions[col_letter].width or sheet.sheet_format.defaultColWidth)

    for row_idx in range(1, sheet.max_row + 1):
        total_height += sheet.row_dimensions[row_idx].height or sheet.sheet_format.defaultRowHeight

    return total_width, total_height
    
